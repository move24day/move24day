# excel_filler.py

import openpyxl
import io
import streamlit as st
import os
import traceback
from datetime import date
import re
import utils # <--- utils 모듈 임포트

try:
    import data
except ImportError:
    st.error("data.py 파일을 찾을 수 없습니다. excel_filler.py와 같은 폴더에 있는지 확인하세요.")
    data = None

# --- 수정된 get_tv_qty (utils 사용) ---
def get_tv_qty(state_data):
    """모든 크기의 TV 수량을 합산하여 반환 (utils.get_item_qty 사용)"""
    if not data or not hasattr(data, 'items') or not isinstance(data.items, dict):
        return 0
    total_tv_qty = 0
    # data.items에서 "TV("로 시작하는 모든 품목 키를 찾습니다.
    tv_keys = [key for key in data.items if key.startswith("TV(")]
    for tv_item_name in tv_keys:
        # utils의 get_item_qty 함수를 사용하여 각 TV 품목의 수량을 가져옵니다.
        total_tv_qty += utils.get_item_qty(state_data, tv_item_name)
    return total_tv_qty
# --- 헬퍼 함수 끝 ---


def fill_final_excel_template(state_data, calculated_cost_items, total_cost, personnel_info):
    """
    final.xlsx 템플릿을 열고 값을 채웁니다.
    경유지 정보 및 요금 포함
    """
    if not data:
        st.error("data.py 모듈 로드 실패로 Excel 생성을 진행할 수 없습니다.")
        return None

    try:
        # final.xlsx 경로 설정
        script_dir = os.path.dirname(__file__) if "__file__" in locals() else "."
        final_xlsx_path = os.path.join(script_dir, "final.xlsx") # 실제 템플릿 파일명

        if not os.path.exists(final_xlsx_path):
            st.error(f"템플릿 파일 '{final_xlsx_path}'을 찾을 수 없습니다.")
            print(f"Error: Template file not found at '{final_xlsx_path}'")
            return None

        wb = openpyxl.load_workbook(final_xlsx_path)
        # 시트 이름 확인 필요 (활성 시트 또는 특정 이름)
        # ws = wb.active # 활성 시트 사용
        ws = wb['Sheet1'] # 또는 특정 시트 이름 사용, 예: 'Sheet1'

        print(f"INFO [Excel Filler]: Template '{final_xlsx_path}' loaded, using sheet '{ws.title}'")

        # --- 1. 기본 정보 입력 ---
        is_storage = state_data.get('is_storage_move', False)
        is_long_distance = state_data.get('apply_long_distance', False)
        has_via_point = state_data.get('has_via_point', False) # 경유지 유무

        move_type_parts = []
        if is_storage: move_type_parts.append("보관")
        if has_via_point: move_type_parts.append("경유") # 경유 추가
        if is_long_distance: move_type_parts.append("장거리")
        
        base_move_type = state_data.get('base_move_type', "")
        if "사무실" in base_move_type: move_type_parts.append("사무실")
        elif "가정" in base_move_type: move_type_parts.append("가정")
        
        move_type_str = " ".join(move_type_parts).strip() or base_move_type
        ws['J1'] = move_type_str

        ws['C2'] = state_data.get('customer_name', '')
        ws['G2'] = state_data.get('customer_phone', '')
        
        moving_date_val = state_data.get('moving_date')
        if isinstance(moving_date_val, date):
            ws['K3'] = moving_date_val
            ws['K3'].number_format = 'yyyy-mm-dd' # 날짜 형식 지정
        elif moving_date_val: # 문자열 등으로 들어올 경우 그대로 사용
            ws['K3'] = str(moving_date_val)
        else:
            ws['K3'] = '' # 값 없을 시 공백

        ws['C3'] = state_data.get('from_location', '')
        ws['C4'] = state_data.get('to_location', '')

        # 경유지 정보 추가 (템플릿에 해당 셀이 있다고 가정, 예: C5)
        if has_via_point:
            ws['G4'] = state_data.get('via_point_location', '') # 예시 셀 'G4', 실제 템플릿에 맞게 수정
        else:
            ws['G4'] = ''


        p_info = personnel_info if isinstance(personnel_info, dict) else {}
        try: ws['L5'] = int(p_info.get('final_men', 0) or 0)
        except (ValueError, TypeError): ws['L5'] = 0
        try: ws['L6'] = int(p_info.get('final_women', 0) or 0)
        except (ValueError, TypeError): ws['L6'] = 0

        from_floor_str = str(state_data.get('from_floor', '')).strip()
        ws['D5'] = f"{from_floor_str}층" if from_floor_str else ''
        to_floor_str = str(state_data.get('to_floor', '')).strip()
        ws['D6'] = f"{to_floor_str}층" if to_floor_str else ''
        
        ws['E5'] = state_data.get('from_method', '')
        ws['E6'] = state_data.get('to_method', '')
        # 경유지 작업 방법 (템플릿에 해당 셀이 있다고 가정, 예: E7)
        if has_via_point:
            ws['K6'] = state_data.get('via_point_method', '') # 예시 셀 'K6'
        else:
            ws['K6'] = ''


        # --- 차량 정보 (B7: 톤수만, H7: 실제 투입) ---
        selected_vehicle = state_data.get('final_selected_vehicle', '')
        vehicle_tonnage = ''
        if isinstance(selected_vehicle, str) and selected_vehicle.strip():
            try:
                match = re.search(r'(\d+(\.\d+)?)', selected_vehicle) # 숫자 부분 추출
                if match:
                    vehicle_tonnage = match.group(1) # "2.5" 또는 "5" 등
                else: # 매칭 실패 시, 숫자 아닌 문자 제거 후 시도
                    vehicle_tonnage_cleaned = re.sub(r'[^\d.]', '', selected_vehicle)
                    vehicle_tonnage = vehicle_tonnage_cleaned if vehicle_tonnage_cleaned else ''
            except Exception as e:
                print(f"ERROR [Excel Filler B7]: Error processing vehicle tonnage: {e}")
                vehicle_tonnage = '' # 오류 시 빈 문자열
        elif selected_vehicle: # 숫자가 아닌 다른 타입일 경우 문자열로 변환
             vehicle_tonnage = str(selected_vehicle)
        ws['B7'] = vehicle_tonnage # "톤" 글자 제외하고 숫자만 입력되도록 수정

        dispatched_parts = []
        dispatched_1t = state_data.get('dispatched_1t', 0)
        dispatched_2_5t = state_data.get('dispatched_2_5t', 0)
        dispatched_3_5t = state_data.get('dispatched_3_5t', 0)
        dispatched_5t = state_data.get('dispatched_5t', 0)
        try: dispatched_1t = int(dispatched_1t or 0)
        except: dispatched_1t = 0
        try: dispatched_2_5t = int(dispatched_2_5t or 0)
        except: dispatched_2_5t = 0
        try: dispatched_3_5t = int(dispatched_3_5t or 0)
        except: dispatched_3_5t = 0
        try: dispatched_5t = int(dispatched_5t or 0)
        except: dispatched_5t = 0
            
        if dispatched_1t > 0: dispatched_parts.append(f"1톤: {dispatched_1t}")
        if dispatched_2_5t > 0: dispatched_parts.append(f"2.5톤: {dispatched_2_5t}")
        if dispatched_3_5t > 0: dispatched_parts.append(f"3.5톤: {dispatched_3_5t}")
        if dispatched_5t > 0: dispatched_parts.append(f"5톤: {dispatched_5t}")
        ws['H7'] = ", ".join(dispatched_parts) if dispatched_parts else ''


        # --- 2. 비용 정보 입력 (경유지 요금 포함) ---
        basic_fare = 0; ladder_from = 0; ladder_to = 0; sky_cost=0; storage_cost=0
        long_dist_cost=0; waste_cost=0; add_person_cost=0; date_surcharge=0
        regional_surcharge=0; adjustment=0; via_point_surcharge = 0 # 경유지 요금 변수

        if calculated_cost_items and isinstance(calculated_cost_items, list):
            for item in calculated_cost_items:
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    label, amount_raw = item[0], item[1]
                    try: amount = int(amount_raw)
                    except (ValueError, TypeError): amount = 0
                    
                    if label == '기본 운임': basic_fare = amount
                    elif label == '출발지 사다리차': ladder_from = amount
                    elif label == '도착지 사다리차': ladder_to = amount
                    elif label == '스카이 장비': sky_cost = amount
                    elif label == '보관료': storage_cost = amount
                    elif label == '장거리 운송료': long_dist_cost = amount
                    elif label == '폐기물 처리(톤)': waste_cost = amount
                    elif label == '추가 인력': add_person_cost = amount
                    elif label == '날짜 할증': date_surcharge = amount
                    elif label == '지방 사다리 추가요금': regional_surcharge = amount
                    elif label == '경유지 추가요금': via_point_surcharge = amount # 경유지 요금 할당
                    elif "조정" in label: adjustment += amount # 할증/할인 조정은 누적

        ws['F22'] = basic_fare
        ws['F23'] = ladder_from + ladder_to # 출발지, 도착지 사다리 합산 (템플릿 구조에 따라 분리 가능)
        ws['J22'] = sky_cost # 스카이 비용 (템플릿 셀 J22 가정)
        # 기타 비용들 (템플릿에 맞는 셀에 배치)
        # 예: ws['X22'] = storage_cost
        # 예: ws['X23'] = long_dist_cost
        # 예: ws['X24'] = waste_cost
        # 예: ws['X25'] = add_person_cost
        # 예: ws['X26'] = date_surcharge
        # 예: ws['X27'] = regional_surcharge
        # 예: ws['X28'] = via_point_surcharge # 경유지 요금 (템플릿 셀 X28 가정)
        # 예: ws['X29'] = adjustment

        # 계약금 및 잔금 (state_manager.py와 키 일관성 확인)
        # UI는 deposit_amount 사용, 저장된 state는 tab3_deposit_amount 일 수 있음
        deposit_amount_raw = state_data.get('deposit_amount', state_data.get('tab3_deposit_amount', 0))
        try: deposit_amount = int(deposit_amount_raw)
        except (ValueError, TypeError): deposit_amount = 0
        ws['J23'] = deposit_amount

        try: total_cost_num = int(total_cost)
        except (ValueError, TypeError): total_cost_num = 0
        ws['F25'] = total_cost_num # 총액
        remaining_balance = total_cost_num - deposit_amount
        ws['J24'] = remaining_balance # 잔금

        # --- 3. 고객 요구사항 입력 (B26 셀부터 순차 기록 - 기존 수정 유지) ---
        special_notes_str = state_data.get('special_notes', '')
        start_row_notes = 26 # 시작 행
        max_possible_note_lines = 20 # 최대 기록 줄 수 (템플릿에 따라 조절)

        # 기존 내용 지우기
        for i in range(max_possible_note_lines):
             clear_cell_addr = f"B{start_row_notes + i}"
             try:
                 if ws[clear_cell_addr].value is not None: # 셀이 존재하고 값이 있을 때만 None으로 설정
                     ws[clear_cell_addr].value = None
             except Exception as e:
                 print(f"Warning [Excel Filler B26+]: Could not clear cell {clear_cell_addr}: {e}")

        if special_notes_str:
            notes_parts = [part.strip() for part in special_notes_str.split('.') if part.strip()] # '.' 기준으로 나누고 공백 제거
            for i, part in enumerate(notes_parts):
                if i < max_possible_note_lines: # 최대 줄 수 넘지 않도록
                    target_cell_notes = f"B{start_row_notes + i}"
                    try:
                        ws[target_cell_notes] = part
                    except Exception as e:
                         print(f"ERROR [Excel Filler B26+]: Failed to write note to {target_cell_notes}: {e}")
        else: # 고객 요구사항이 없을 경우, 첫 줄만 비움 (이미 위에서 처리됨)
             try:
                 if ws['B26'].value is not None: ws['B26'] = None
             except Exception as e: print(f"Warning [Excel Filler B26+]: Could not clear B26 for empty notes: {e}")


        # --- 4. 품목 수량 입력 (utils.get_item_qty 사용, D8 장롱 수량 처리) ---
        # D열
        original_jangrong_qty = utils.get_item_qty(state_data, '장롱') # utils 사용
        jangrong_formatted_qty = "0.0" # 기본 문자열 값
        try:
            # 장롱은 3으로 나눈 값을 소수점 첫째 자리까지 표시 (예: 10자 -> 3.3)
            calculated_qty = original_jangrong_qty / 3.0
            jangrong_formatted_qty = f"{calculated_qty:.1f}"
        except ZeroDivisionError: # 0으로 나누는 경우 (거의 발생 안 함)
            jangrong_formatted_qty = "0.0"
        except Exception as e:
            print(f"ERROR [Excel Filler D8]: Error calculating Jangrong qty: {e}")
            jangrong_formatted_qty = "Error" # 오류 발생 시 "Error" 표시
        ws['D8'] = jangrong_formatted_qty # 계산된 값 또는 오류 메시지 입력

        ws['D9'] = utils.get_item_qty(state_data, '더블침대')
        ws['D10'] = utils.get_item_qty(state_data, '서랍장')
        ws['D11'] = utils.get_item_qty(state_data, '서랍장(3단)')
        ws['D12'] = utils.get_item_qty(state_data, '4도어 냉장고')
        ws['D13'] = utils.get_item_qty(state_data, '김치냉장고(일반형)')
        ws['D14'] = utils.get_item_qty(state_data, '김치냉장고(스탠드형)')
        ws['D15'] = utils.get_item_qty(state_data, '소파(3인용)')
        ws['D16'] = utils.get_item_qty(state_data, '소파(1인용)')
        ws['D17'] = utils.get_item_qty(state_data, '식탁(4인)')
        ws['D18'] = utils.get_item_qty(state_data, '에어컨')
        ws['D19'] = utils.get_item_qty(state_data, '장식장')
        ws['D20'] = utils.get_item_qty(state_data, '피아노(디지털)')
        ws['D21'] = utils.get_item_qty(state_data, '세탁기 및 건조기')

        # H열
        ws['H9'] = utils.get_item_qty(state_data, '사무실책상')
        ws['H10'] = utils.get_item_qty(state_data, '책상&의자')
        ws['H11'] = utils.get_item_qty(state_data, '책장')
        ws['H15'] = utils.get_item_qty(state_data, '바구니')
        ws['H16'] = utils.get_item_qty(state_data, '중박스') # data.py 정의에 따라 '중자바구니' 또는 '중박스' 확인
        ws['H19'] = utils.get_item_qty(state_data, '화분')
        ws['H20'] = utils.get_item_qty(state_data, '책바구니')

        # L열
        ws['L8'] = utils.get_item_qty(state_data, '스타일러')
        ws['L9'] = utils.get_item_qty(state_data, '안마기')
        ws['L10'] = utils.get_item_qty(state_data, '피아노(일반)')
        ws['L12'] = get_tv_qty(state_data) # 수정된 get_tv_qty 호출 (모든 TV 합산)
        ws['L16'] = utils.get_item_qty(state_data, '금고')
        ws['L17'] = utils.get_item_qty(state_data, '앵글')


        # --- 5. 완료된 엑셀 파일을 메모리에 저장 ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0) # 버퍼 포인터 리셋
        print("INFO [Excel Filler]: Excel file generation complete.")
        return output.getvalue() # 바이트 데이터 반환

    except FileNotFoundError:
        st.error(f"Excel 템플릿 파일 '{final_xlsx_path}'을(를) 찾을 수 없습니다.")
        print(f"Error: Template file not found at '{final_xlsx_path}' during generation.")
        return None
    except Exception as e:
        st.error(f"Excel 생성 중 오류 발생: {e}")
        print(f"Error during Excel generation: {e}")
        traceback.print_exc() # 콘솔/로그에 상세 오류 출력
        return None